import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from datetime import datetime

# base: abrir arquivo Excel
arquivo = openpyxl.load_workbook('Treinos/Code_trabalho/BASE DESCARGAS-2026 .xlsx', data_only=True)
abas_do_arquivo = arquivo.sheetnames

# base: mostrar planilhas disponíveis e pedir escolha do usuário
for i, aba in enumerate(abas_do_arquivo):
    print(f'{i+1}: {aba}')

aba_escolhida = int(input('Escolha o número da aba: '))
planilha_escolhida = arquivo[abas_do_arquivo[aba_escolhida - 1]]

# base: tornar a planilha escolhida disponível e obter o cabeçalho
cabecalho = [cell.value for cell in planilha_escolhida[2]]
print("Cabeçalho:", cabecalho)

# tratamento: verificar se colunas obrigatórias existem e estão na posição correta
colunas_obrigatorias = {"DT": 0, "EIXO": 4, "valor ": 23}
for coluna, posicao in colunas_obrigatorias.items():
    if coluna not in cabecalho:
        raise ValueError(f"Erro: coluna '{coluna}' não encontrada no cabeçalho.")
    elif cabecalho.index(coluna) != posicao:
        raise ValueError(f"Erro: coluna '{coluna}' está na posição {cabecalho.index(coluna)}, mas deveria estar em {posicao}.")

# base: abrir navegador uma vez
navegador = webdriver.Chrome()
wait = WebDriverWait(navegador, 10)

# tratamento: carregar log de DTs já enviadas para evitar reenvio
# agora o log será organizado em pastas por ano/mês
data_atual = datetime.now()
ano = data_atual.strftime("%Y")
mes = data_atual.strftime("%B")  # mês por extenso em inglês
dia = data_atual.strftime("%d/%m/%Y")

# cria estrutura de pastas: enviados/ano/mês
pasta_base = "enviados"
pasta_ano = os.path.join(pasta_base, ano)
pasta_mes = os.path.join(pasta_ano, mes)

os.makedirs(pasta_mes, exist_ok=True)

# arquivo de log dentro da pasta do mês
arquivo_log = os.path.join(pasta_mes, "log.txt")

try:
    with open(arquivo_log, "r") as f:
        enviados = set(f.read().splitlines())
except FileNotFoundError:
    enviados = set()

# tratamento 3: capturar encerramento prematuro da automação
try:
    # base: identificar última linha e ler colunas necessárias
    ultima_linha = planilha_escolhida.max_row
    for i, linha in enumerate(planilha_escolhida.iter_rows(min_row=3, max_row=ultima_linha)):
        dt = linha[0].value   # coluna A → DT
        eixo = linha[4].value # coluna E → EIXO
        valor = linha[23].value # coluna X → VALOR

        # tratamento: interromper se faltar dados obrigatórios
        if dt is None or eixo is None or valor is None:
            print(f"Interrompendo na linha {i+3} por falta de dados.")
            break

        # tratamento: evitar reenvio de DT já processada
        if str(dt) in enviados:
            print(f"DT {dt} já enviada anteriormente. Pulando...")
            continue

        print(f'{i+1}: DT: {dt}, EIXO: {eixo}, VALOR: {valor}')

        # base: acessar formulário
        time.sleep(3)
        navegador.get('https://forms.office.com/r/cBcVBYa5mZ')

        # optimização: usar WebDriverWait em vez de time.sleep
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="question-list"]/div[1]/div[2]/div/span/input'))).send_keys("JSL")
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="question-list"]/div[2]/div[2]/div/span/input'))).send_keys("Operacaosuzano@jsl.com.br")
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="question-list"]/div[3]/div[2]/div/span/input'))).send_keys(str(dt))
        
        navegador.find_element(By.XPATH, '//*[@id="question-list"]/div[4]/div[2]/div/div/div[2]/div/label/span[1]/input').click()
        
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="question-list"]/div[5]/div[2]/div/span/input'))).send_keys(str(eixo))
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="question-list"]/div[6]/div[2]/div/span/input'))).send_keys(str(valor))
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="question-list"]/div[7]/div[2]/div/span/input'))).send_keys("0")
        
        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="form-main-content1"]/div/div/div[2]/div[3]/div/button'))).click()

        # tratamento: registrar DT enviada no log com data completa
        with open(arquivo_log, "a") as f:
            f.write(f"DT {dt} enviado em {dia}\n")

except KeyboardInterrupt:
    # tratamento 3: mensagem clara ao encerrar manualmente
    print("\nAutomação encerrada pelo usuário.")

finally:
    # base: fechar navegador ao final
    navegador.quit()

