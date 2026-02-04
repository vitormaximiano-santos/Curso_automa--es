import openpyxl as px 

# Carregar uma planilha existente
arquivo = px.load_workbook("Openpyxl\Aula_03\planilha_funcionarios.xlsx")

# Selecionar a planilha pelo nome, da o nome "funcionarios" para a planilha
planilha_funcionarios = arquivo['Funcionários']

# Imprimir o objeto da planilha selecionada
print(planilha_funcionarios)

# Acessar o valor da célula D6 (salário do Antonio)
salario_antonio = planilha_funcionarios['D6'].value
print(f'Salário do Antonio: R$ {salario_antonio:.2f}')


linha_13 = planilha_funcionarios[13]

# Iterar sobre cada célula na linha 13
for celula in linha_13:
    
    # Imprimir o valor de cada célula na linha 13
    print(celula.value)
    
    
# pegando varia linhas de uma vez, por exemplo, da linha 7 até a linha 12
linhas_7_a_12 = planilha_funcionarios[7:13] # o 13 é exclusivo, ou seja, vai até a linha 12

# Iterar sobre cada linha do intervalo selecionado , itera = passar por cada um
for linha in linhas_7_a_12:
    print ('-' * 50)
    
    # Iterar sobre cada célula na linha
    for celula in linha:
        
        # Imprimir o valor de cada célula na linha
        print(celula.value)
        
        
# pegando uma coluna inteira, por exemplo, a coluna dos salarios (coluna D)
coluna_salarios = planilha_funcionarios['D']
print(coluna_salarios)


# Iterar sobre cada célula na planilha, pegando apenas os valores das células
for linha in planilha_funcionarios.iter_rows(values_only=True):
    print ('-' * 50)
    
    # Iterar sobre cada célula na linha
    for celula in linha:
        print(celula)
        
        
# Iterar sobre cada linha na planilha, pegando apenas os valores das células e pulando a primeira linha (cabeçalho)      
for linha in planilha_funcionarios.iter_rows(values_only=True,min_row=2):
    print ('-' * 50)

    # Desempacotar os valores da linha em variáveis
    nome, departamento, idade, salario, data_admissao = linha
    
    print(f"""Nome: {nome}
Departamento: {departamento} ,
Idade: {idade} ,
Salário: R$ {salario:.2f} ,
Data de Admissão: {data_admissao.strftime('%d/%m/%Y')}  
""")
    
    # data é um objeto do tipo datetime, por isso usamos o strftime para formatar a data para o formato dia/mês/ano