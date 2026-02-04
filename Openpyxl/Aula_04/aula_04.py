import openpyxl as px
import openpyxl.styles as pxs
import datetime

# Abrindo o arquivo
arquivo = px.Workbook()

# Selecionando a planilha
planilha = arquivo.active

# Adicionando um título
planilha.title = 'Produtos'

# criando objeto de data, "datetime.now()" retorna a data atual
agora = datetime.datetime.now()

# Adicionando colunas
planilha.append(['Produto', 'Preço', 'Quantidade', 'Data'])

# Adicionando dados
planilha.append(['Camiseta', 60.00  , 10,agora])
planilha.append(['Calça Jeans', 120.00  , 5,agora])
planilha.append(['Tenis', 250.00  , 2,agora])

# Criando título para a tabela
planilha.merge_cells('A1:D1')
planilha["A1"] = "Relatório de Vendas" # título da planilha
planilha["A1"].alignment = pxs.Alignment(horizontal="center") # alinhamento central
planilha["A1"].font = px.styles.Font(bold=True, size=14, color="FFFFFF") # cor branca
planilha["A1"].fill = px.styles.PatternFill(fgColor="1f497D", fill_type="solid") # fundo na cor azul claro

# criando o estilo da fonte dos cabeçalhos, bold simboliza que o texto será em negrito e size define o tamanho da fonte, na cor #720583 que é roxo, sendo o valor hexadecimal da cor
config_fonte_cabecalho = px.styles.Font(bold=True, size=10, color="720583")

# criando o estilo do fundo dos cabeçalhos, sendo "fgColor" a cor de primeiro plano e "fill_type" o tipo de preenchimento
confid_fundo_cabecalho = px.styles.PatternFill(fgColor="4f4f4f", fill_type="solid")

# configurando a borda
fina = pxs.Side(style="thin") # estilo de borda
borda = pxs.Border(left=fina, right=fina, top=fina, bottom=fina) # criando objeto de borda


# Aplicando o estilo da fonte dos cabeçalhos, maneira mais rápida, usando loops
for celula in planilha[2]:
    
    # aplicando o estilo da fonte dos cabeçalhos, ".font" é um atributo de objeto que pode ser aplicado a qualquer objeto do openpyxl, o que permite aplicar estilos a objetos específicos.
    celula.font = config_fonte_cabecalho
    
    # aplicando o estilo de preenchimento, "fill" é um atributo de objeto que pode ser aplicado a qualquer objeto do openpyxl, o que permite aplicar estilos a objetos específicos.
    celula.fill = confid_fundo_cabecalho
    
    # diferença entre "fill" e "font" é que "fill" aplica o preenchimento apenas no fundo da célula, enquanto "font" aplica o preenchimento a fonte.

# Aplicando o estilo da fonte dos cabeçalhos, ".font" é um atributo de objeto que pode ser aplicado a qualquer objeto do openpyxl, o que permite aplicar estilos a objetos específicos, forma mais lenta
#planilha['A1'].font = config_fonte_cabecalho
#planilha['B1'].font = config_fonte_cabecalho

# iterando sobre as células para aplicar o alinhamento à célula
for celula in planilha["B"]:
    
    # ignorando o cabeçalho
    if celula.row <= 2: 
        continue
        
    # aplicando alinhamento à célula
    celula.alignment = pxs.Alignment(horizontal="right")
    
    # aplicando formato de número à célula, os "#,##0.00" representam o formato de número, sendo que "#" é o número de casas decimais, ",##0.00" é o formato de número
    celula.number_format = "R$ #,##0.00"


# iterando sobre as linhas para aplicar a borda à célula
for linha in planilha.iter_rows():
    
    # iterando sobre as células para aplicar a borda à célula
    for celula in linha:
        
        # aplicando bordas à célula
        celula.border = borda

# iterando sobre as células para aplicar o formato de data à célula
for celuna in planilha['D']:
    
    # aplicando formato de data à célula
    celuna.number_format = "dd/mm/yyyy" # "dd" representa dia, "mm" representa mês, "yyyy" representa ano

# salvando o arquivo
arquivo.save('Openpyxl\Aula_04\produtos.xlsx')