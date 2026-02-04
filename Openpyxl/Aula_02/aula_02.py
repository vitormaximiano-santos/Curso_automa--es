import openpyxl as px

# Criando uma planilha nova
arquivo = px.Workbook()

# selecionando a planilha ativa
planilha_atual = arquivo.active

# renomeando a planilha
planilha_atual.title = "Produtos"

# adicionando dados na planilha
# cabeçalho
planilha_atual["A1"] = "Nome"
planilha_atual["B1"] = "Preço"

# primeira linha de dados
planilha_atual["A2"] = "Caneta"
planilha_atual["B2"] = 1.50

# adicionando uma linha de dados,usando lista   
planilha_atual.append(["Caderno", 15.00])

# alterando o preço do caderno
planilha_atual["b3"] = 20.00  

# criando uma nova planilha/aba
planilha_vendas = arquivo.create_sheet("Vendas")

# adicionando dados na nova planilha: cabeçalho
planilha_vendas.append(['valor','data'])

# adicionando uma linha de dados
planilha_vendas.append([150.00,'01/06/2024'])

# mostrando os nomes das planilhas/abas
print(arquivo.sheetnames)

# salvando a planilha
arquivo.save("Openpyxl/Aula_02/planilhas.xlsx")

