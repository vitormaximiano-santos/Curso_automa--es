import openpyxl as px
import openpyxl.styles as pxs

# Abrindo o arquivo
arquivo = px.Workbook()

# Selecionando a planilha
planilha = arquivo.active

# nome da planilha
planilha.title = 'Livros'

# criando um titulo para a tabela
planilha.merge_cells('A1:D1') # mesclando as células A1:D1
planilha["A1"] = "Diário de Leituras – Agosto 2025" # título da planilha
planilha["A1"].alignment = pxs.Alignment(horizontal="center") # alinhamento central
planilha["A1"].fill = px.styles.PatternFill(fgColor="1f497D", fill_type="solid") # fundo na cor azul claro
planilha["A1"].font = px.styles.Font(bold=True, size=14, color="FFFFFF") # cor branca

# criando colunas
planilha.append(['Livro', 'Autor', 'Data de Início', 'Progresso (%)'])

# criando registros fictícios de livros/ adicionando dados
planilha.append(['Python para Iniciantes', 'Fernando M. Perez', '20/08/2025', 10.0])
planilha.append(['Aprenda Python Básico', 'Fernando M. Perez', '20/08/2025', 10.0])
planilha.append(['Aprenda Python Avançado', 'Fernando M. Perez', '20/08/2025', 10.0])

# iterando sobre as células para aplicar a formatação
for celula in planilha[1]:
    # aplicando formatação
    celula.font = pxs.Font(color="FFFFFF", bold=True, size=10) # cor branca, negrito, tamanho 10
    celula.fill = pxs.PatternFill(fgColor="4f4f4f", fill_type="solid") # fundo cinza escuro
    
    
for linha in planilha.iter_rows():
    for celula in linha:
        celula.border = pxs.Border(pxs.Side(style="thin"),pxs.Side(style="thin"),pxs.Side(style="thin"),pxs.Side(style="thin")) # bordas finas
        
        if celula.row <= 2:
            continue
        
        # identificando se a linha é par ou ímpar, para aplicar a cor de fundo
        if celula.row % 2 == 0:
            celula.fill = pxs.PatternFill(fgColor="C9C9C9", fill_type="solid") # fundo cinza claro
            
            
for celula in planilha['C']:
    celula.number_format = "dd/mm/yyyy" # formato de data
    
for celula in planilha['D']:
    celula.number_format = "0.0%" # formato de percentagem
    
# salvando o arquivo
arquivo.save('Openpyxl\Exercicios\Ex_aula_04\diario_leituras.xlsx')