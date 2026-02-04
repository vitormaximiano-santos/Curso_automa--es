import openpyxl as px
import shutil

# Descompacta o arquivo ZIP
shutil.unpack_archive('Openpyxl/Aula_01/Excel_aula_01.zip', 'Openpyxl/Aula_01/Excel_aula_01.zip_descompactado', 'zip')

# Carrega a planilha Excel
arquivo = px.load_workbook('Openpyxl/Aula_01/Excel_aula_01.zip_descompactado/Alunos.xlsx')

# Exibe os nomes das planilhas/abas presentes no arquivo Excel
print(arquivo.sheetnames)  

# Seleciona a aba ativa (primeira aba por padrão)
aba_ativa = arquivo.active

# Exibe o título da aba ativa
print(aba_ativa.title) 

# Seleciona uma aba específica pelo nome
aba_alunos = arquivo['Planilha1']

# Exibe o conteúdo da aba selecionada
print(aba_alunos)

# Exibe o valor da célula A1 da aba selecionada,value faz a leitura do conteudo da celula
valor_a1 = aba_alunos['A1'].value
print()

# outra forma de acessar o valor de uma célula
valor_b1 = aba_alunos.cell(row=1, column=2).value
print(valor_b1)  

# Modifica o valor da célula B1
aba_alunos.cell(row=1, column=2).value = "prova 1"
valor_b1 = aba_alunos.cell(row=1, column=2).value
print(valor_b1)


# cria uma copia do arquivo excel com as modificações
arquivo.save('Openpyxl/Aula_01/Excel_aula_01_modificado.xlsx')

# substitui o valor no arquivo original
#arquivo.save('Openpyxl/Aula_01/Excel_aula_01.zip_descompactado/Alunos.xlsx')

# descobrir a ultima linha, 1º forma
ultima_linha = aba_alunos.max_row # Retorna o número da última linha preenchida

print(f"A ultima linha da planilha é: {ultima_linha}")

# descobrir a ultima linha, 2º forma
print(len(aba_alunos['A']))  # Conta o número de células preenchidas na coluna A