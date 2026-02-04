# importar a biblioteca openpyxl para manipulação de planilhas Excel
import openpyxl as px

# importar a biblioteca copy para copiar estilos de células
import copy

# Carregar a planilha existente
bairros = px.load_workbook('Openpyxl/Aula_01/Excel_aula_01.zip_descompactado/Bairros.xlsx')

# Exibe os nomes das planilhas/abas na planilha carregada
print(bairros.sheetnames)  

# Seleciona a planilha "Base de Dados"
aba_dados = bairros['Base de Dados']

# Copia o estilo do cabeçalho da planilha de dados
estilo_cabecalho = copy.copy(aba_dados['A1']._style)

# Obtém o número da última linha com dados na planilha selecionada
ultima_linha = aba_dados.max_row
print(f'A última linha da planilha é: {ultima_linha}')

# Função para criar uma nova aba com o nome do bairro
def criar_aba(bairro, bairros, estilo_cabecalho):
    
    # Verifica se a aba com o nome do bairro já existe
    if bairro not in bairros.sheetnames:
        
        # criar uma nova aba com o nome do bairro
        bairros.create_sheet(title=bairro) 
        
        # Seleciona a nova aba criada
        nova_aba = bairros[bairro]  
        
        # Adiciona os cabeçalhos nas colunas da nova aba
        nova_aba["A1"].value = "Data de Nascimento" # Cabeçalho da coluna A
        nova_aba["B1"].value = "Nome" # Cabeçalho da coluna B
        nova_aba["C1"].value = "Bairro" # Cabeçalho da coluna C
        
        # Aplica o estilo do cabeçalho da aba de dados à nova aba
        nova_aba["A1"]._style = estilo_cabecalho # Cabeçalho da coluna A
        nova_aba["B1"]._style = estilo_cabecalho # Cabeçalho da coluna B
        nova_aba["C1"]._style = estilo_cabecalho # Cabeçalho da coluna C
        
# Função para transferir informações de uma aba para outra
def transferir_informacoes_aba(aba_origem, aba_destino, linha_origem):
    # Obtém os valores das células da linha de origem
    linha_destino = aba_destino.max_row + 1  # Próxima linha disponível na aba de destino

    # Loop para copiar os valores das colunas A, B e C (1, 2 e 3)
    for coluna in range(1,4):
        # Copia o valor da célula da aba de origem para a aba de destino
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna)
        
        # Define a célula de destino na aba de destino
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
        
        # Atribui o valor da célula de origem para a célula de destino
        celula_destino.value = celula_origem.value
        
        # Copia o estilo da célula de origem para a célula de destino
        celula_destino._style = copy.copy(celula_origem._style)
        


# Loop para percorrer as linhas da planilha de dados, começando da linha 2 até a última linha
for linha in range(2, ultima_linha + 1):
    
    # Obtém o valor da célula na coluna 3 (Bairro) para a linha atual (1º forma)
    bairro = aba_dados.cell(row = linha, column = 3).value
    
      # 2º forma , diferente da anterior pois usa f-string.
    # bairro = aba_dados[f'C{linha}'].value  
    
    # Verifica se o valor do bairro é vazio ou None
    if not bairro:
        break
    
    # criar uma nova aba para cada bairro
    criar_aba (bairro, bairros, estilo_cabecalho)
    
    aba_destino = bairros[bairro]
    transferir_informacoes_aba(aba_dados, aba_destino, linha)
    
# Salvar as alterações na planilha
bairros.save('Openpyxl/Aula_01/Bairros_modificado.xlsx')